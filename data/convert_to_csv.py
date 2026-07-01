"""
One-time script to convert Excel product data to CSV files.
Run from the project root: python data/convert_to_csv.py

Outputs (gitignored via data/):
  data/csv/products.csv       — product catalog from references product_ecommerce.xlsx
  data/csv/retail_prices.csv  — retail pricing from DEALER PRICES 2025.xlsx
"""

import csv
import datetime
import os
import re
import subprocess
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_DIR = os.path.join(BASE_DIR, "csv")
os.makedirs(CSV_DIR, exist_ok=True)

RAW_DIR = os.path.join(BASE_DIR, "raw")

PRODUCTS_CSV = os.path.join(CSV_DIR, "products.csv")
PRODUCT_STUBS_CSV = os.path.join(CSV_DIR, "product_stubs.csv")
RETAIL_PRICES_CSV = os.path.join(CSV_DIR, "retail_prices.csv")
DEALER_PRICES_CSV = os.path.join(CSV_DIR, "dealer_prices.csv")
MERGED_IMPORT_CSV = os.path.join(CSV_DIR, "import_all_merged.csv")
COMPATIBILITY_OPTIONS_CSV = os.path.join(CSV_DIR, "compatibility_options.csv")
VISIBLE_CATEGORIES_TXT = os.path.join(RAW_DIR, "visible_categories.txt")

REFERENCE_PATTERN = re.compile(r"(?<!\d)(\d{2}\.\d{3}\.\d{3}\.\d{2}-\d)(?!\d)")


def resolve_source_file(filename):
    """Resolve source files from data/raw first, then fallback to data/."""
    candidates = [
        os.path.join(RAW_DIR, filename),
        os.path.join(BASE_DIR, filename),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    raise FileNotFoundError(f"Source file not found: {filename}")


def format_reference(num):
    """Convert numeric reference 49604000082 → '49.604.000.08-2'"""
    s = str(int(num))
    if len(s) == 11:
        return f"{s[0:2]}.{s[2:5]}.{s[5:8]}.{s[8:10]}-{s[10]}"
    return s


def normalize_ref(ref_str):
    """Strip all non-alphanumeric chars and lowercase for matching."""
    return re.sub(r"[^a-z0-9]", "", str(ref_str).lower())


def ref_to_regex(ref_str):
    """Convert wildcard reference to a regex, matching x/y wildcard groups."""
    normalized = normalize_ref(ref_str)
    pattern = re.sub(
        r"[xy]+", lambda m: r"\\d{" + str(len(m.group())) + r"}", normalized
    )
    return re.compile(r"^" + pattern + r"$")


def parse_reference_parts(reference):
    """Split reference into stable numeric groups for easier option handling."""
    m = re.match(r"^(\d{2})\.(\d{3})\.(\d{3})\.(\d{2})-(\d)$", str(reference).strip())
    if not m:
        return {
            "segment_1": "",
            "segment_2": "",
            "segment_3": "",
            "segment_4": "",
            "check_digit": "",
        }
    return {
        "segment_1": m.group(1),
        "segment_2": m.group(2),
        "segment_3": m.group(3),
        "segment_4": m.group(4),
        "check_digit": m.group(5),
    }


def normalize_system_name(name):
    """Normalize system names for resilient matching."""
    return re.sub(r"[^A-Z0-9]", "", str(name).upper())


def load_active_categories(categories_path):
    """Load allowed/active system categories from categories.txt."""
    if not os.path.exists(categories_path):
        return {}

    with open(categories_path, encoding="utf-8") as f:
        raw = f.read()

    tokens = []
    for part in re.split(r"[,\n]+", raw):
        cleaned = part.strip()
        if cleaned:
            tokens.append(cleaned)

    by_norm = {}
    for token in tokens:
        by_norm[normalize_system_name(token)] = token
    return by_norm


def convert_products():
    path = resolve_source_file("references product_ecommerce.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    out_path = os.path.join(CSV_DIR, "products.csv")
    count = 0
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["reference", "reference_num", "name", "ean13"])
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            ref_num, name, _title, ean13 = row[0], row[1], row[2], row[3]
            if not ref_num or not name:
                continue
            ref_formatted = format_reference(ref_num)
            writer.writerow([ref_formatted, int(ref_num), name.strip(), ean13 or ""])
            count += 1

    print(f"products.csv: {count} rows → {out_path}")


def convert_dealer_prices():
    path = resolve_source_file("DEALER PRICES 2025.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Dealer 2025"]

    count = 0
    current_section = ""

    with open(DEALER_PRICES_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["section", "name", "detail", "reference", "price_eur"])

        for row in ws.iter_rows(values_only=True):
            col_a, col_b, col_c, col_d = row[0], row[1], row[2], row[3]

            if all(v is None for v in [col_a, col_b, col_c, col_d]):
                continue
            if col_a in ("DEALER PRICES 2026", "DEALER PRICES 2025"):
                continue
            if isinstance(col_d, datetime.datetime) or isinstance(
                col_c, datetime.datetime
            ):
                continue

            # Section header: col_a has a name but no numeric price (check before skipping "Item reference")
            if (
                col_a
                and col_c in (None, "Item reference")
                and not isinstance(col_d, (int, float))
            ):
                current_section = str(col_a).strip()
                continue

            if col_c and isinstance(col_d, (int, float)):
                name = str(col_a).strip() if col_a else ""
                detail = str(col_b).strip() if col_b else ""
                reference = str(col_c).strip()
                price = float(col_d)
                writer.writerow([current_section, name, detail, reference, price])
                count += 1

    print(f"dealer_prices.csv: {count} rows → {DEALER_PRICES_CSV}")


def convert_retail_prices():
    path = resolve_source_file("DEALER PRICES 2025.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Retail 2025"]

    out_path = os.path.join(CSV_DIR, "retail_prices.csv")
    count = 0
    current_section = ""

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["section", "name", "detail", "reference", "price_eur"])

        for row in ws.iter_rows(values_only=True):
            col_a, col_b, col_c, col_d = row[0], row[1], row[2], row[3]

            # Skip completely empty rows and header rows
            if all(v is None for v in [col_a, col_b, col_c, col_d]):
                continue
            if col_a in ("PRICES 2025",):
                continue
            if isinstance(col_d, datetime.datetime):
                continue

            # Section header: col_a has a name but no usable price data
            if (
                col_a
                and col_c in (None, "Item reference")
                and not isinstance(col_d, (int, float))
            ):
                current_section = str(col_a).strip()
                continue

            if col_c in ("Item reference",):
                continue

            # Product row: has reference in col_c
            if col_c:
                name = str(col_a).strip() if col_a else ""
                detail = str(col_b).strip() if col_b else ""
                reference = str(col_c).strip()
                price = float(col_d) if col_d is not None else None

                # Skip rows with no name and no price (just extra references for same item)
                if not name and price is None:
                    continue

                writer.writerow(
                    [
                        current_section,
                        name,
                        detail,
                        reference,
                        price if price is not None else "",
                    ]
                )
                count += 1
            elif col_d is not None:
                # Rows with price but no reference (e.g. Novox Pack)
                name = str(col_a).strip() if col_a else ""
                writer.writerow([current_section, name, "", "", float(col_d)])
                count += 1

    print(f"retail_prices.csv: {count} rows → {out_path}")


def load_pdf_text(pdf_path):
    """Extract catalog text with pdftotext first; fallback to pypdf."""
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", pdf_path, "-"],
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout
    except (FileNotFoundError, subprocess.CalledProcessError):
        try:
            import pypdf

            reader = pypdf.PdfReader(pdf_path)
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception:
            return ""


def extract_option_tokens(line_text, section="", ch_vals=None):
    """Extract labeled key:value option tokens from a catalog line.

    Returns pipe-separated pairs like ``αS:43º|αC:29º|GH(mm):0.3``.
    For DYNAMIC 3TIBASE sections, ``ch_vals`` (e.g. ["5","7","9"]) produces
    ``αS(CH=5mm):25º|αS(CH=7mm):20º|αS(CH=9mm):10º`` labels instead.
    """
    cleaned = REFERENCE_PATTERN.sub(" ", line_text)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return ""

    letter_tokens = re.findall(r"\b[A-E]\b", cleaned)
    angle_tokens = re.findall(r"\b\d{1,2}º\b", cleaned)
    number_tokens = re.findall(r"\b\d+(?:[\.,]\d+)?\b", cleaned)

    seen: set = set()
    unique_angles = []
    for t in angle_tokens:
        if t not in seen:
            seen.add(t)
            unique_angles.append(t)

    angle_nums = {t.rstrip("º°") for t in unique_angles}
    unique_letters = []
    for t in letter_tokens:
        if t not in seen:
            seen.add(t)
            unique_letters.append(t)

    unique_numbers = []
    for t in number_tokens:
        if t not in seen and t not in angle_nums and not re.fullmatch(r"\d{4}", t):
            seen.add(t)
            unique_numbers.append(t)

    sec_upper = section.upper() if section else ""
    is_3tibase = "3TIBASE" in sec_upper
    is_screw = "SCREW" in sec_upper
    is_scanbody = "SCANBODY" in sec_upper and "OP" not in sec_upper
    is_scanbody_op = "SCANBODY OP" in sec_upper

    labeled = []
    if is_3tibase and ch_vals:
        for i, at in enumerate(unique_angles[: len(ch_vals)]):
            labeled.append(f"αS(CH={ch_vals[i]}mm):{at}")
    else:
        angle_field_names = ["αS", "αC", "αdi"]
        for i, at in enumerate(unique_angles[:3]):
            name = angle_field_names[i] if i < len(angle_field_names) else f"α{i + 1}"
            labeled.append(f"{name}:{at}")
    for lt in unique_letters[:2]:
        labeled.append(f"Typ:{lt}")
    for nt in unique_numbers[:3]:
        val = nt.replace(",", ".")
        if is_screw:
            labeled.append(f"L(mm):{val}")
        elif is_scanbody or is_scanbody_op:
            labeled.append(f"H(mm):{val}")
        else:
            labeled.append(f"GH(mm):{val}")

    return "|".join(labeled[:8])


def parse_pdf_compatibility_rows(pdf_text):
    """Parse rows with references, compatibility code and option tokens from catalog text."""
    rows = []
    current_code = ""
    current_section = ""
    current_ch_vals: list = []
    expect_code_next_line = False
    current_engaging: int | None = (
        None  # 0=non-engaging, 1=engaging (single-column sections)
    )
    engaging_col_positions = (
        None  # (ne_col, e_col) char positions for two-column tables
    )

    for raw_line in pdf_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        compatible = re.search(
            r"COMPATIBLE\s+WITH\s*([0-9]{4}[A-Z]?)", line, flags=re.IGNORECASE
        )
        if compatible:
            current_code = compatible.group(1).upper()
            continue

        if re.fullmatch(r"COMPATIBLE\s+WITH", line, flags=re.IGNORECASE):
            expect_code_next_line = True
            continue

        if expect_code_next_line and re.fullmatch(
            r"\d{4}[A-Z]?", line, flags=re.IGNORECASE
        ):
            current_code = line.upper()
            expect_code_next_line = False
            continue

        expect_code_next_line = False

        # Section header: allow digits so "DYNAMIC 3TIBASE" is captured correctly.
        if (
            re.fullmatch(r"[A-Z][A-Z0-9\s\-/()]{2,60}", line)
            and "PRODUCT REFERENCES" not in line
        ):
            current_section = line
            current_ch_vals = []
            current_engaging = None
            engaging_col_positions = None

        # CH= header line for DYNAMIC 3TIBASE (e.g. "CH=5mm   CH=7mm   CH=9mm")
        ch_header = re.findall(r"CH=(\d+)mm", line, re.IGNORECASE)
        if ch_header and not REFERENCE_PATTERN.search(line):
            current_ch_vals = ch_header
            continue

        # ENGAGING / NON ENGAGING header detection (uses raw_line to preserve column positions)
        if "ENGAGING" in raw_line.upper() and not REFERENCE_PATTERN.search(raw_line):
            ne_m = re.search(r"NON\s+ENGAGING", raw_line, re.IGNORECASE)
            e_candidates = list(re.finditer(r"\bENGAGING\b", raw_line, re.IGNORECASE))
            e_m = next(
                (m for m in e_candidates if ne_m is None or m.start() > ne_m.end()),
                None,
            )
            if ne_m and e_m:
                # Two-column table: NON ENGAGING left, ENGAGING right
                engaging_col_positions = (ne_m.start(), e_m.start())
            elif ne_m:
                current_engaging = 0
                engaging_col_positions = None
            elif e_m:
                current_engaging = 1
                engaging_col_positions = None
            continue

        refs = REFERENCE_PATTERN.findall(line)
        if not refs:
            continue
        if not current_code:
            continue

        option_tokens = extract_option_tokens(line, current_section, current_ch_vals)
        for ref in refs:
            # Determine engaging: use column position for two-column tables, else current state
            if engaging_col_positions is not None:
                ne_col, e_col = engaging_col_positions
                ref_m = re.search(re.escape(ref), raw_line)
                if ref_m:
                    ref_mid = (ref_m.start() + ref_m.end()) // 2
                    engaging = 0 if abs(ref_mid - ne_col) <= abs(ref_mid - e_col) else 1
                else:
                    engaging = current_engaging
            else:
                engaging = current_engaging

            parts = parse_reference_parts(ref)
            rows.append(
                {
                    "compatibility_code": current_code,
                    "section": current_section,
                    "reference": ref,
                    "reference_prefix": parts["segment_1"],
                    "reference_group": parts["segment_2"],
                    "reference_family": parts["segment_3"],
                    "reference_option": parts["segment_4"],
                    "reference_variant": parts["check_digit"],
                    "options": option_tokens,
                    "source_line": line,
                    "engaging": engaging,
                }
            )

        if (
            current_code == "0268"
            and "SCREW" in current_section
            and "LENGTH" in current_section
            and "43.632.201.01-2" in refs
        ):
            current_code = ""
            current_section = ""
            current_ch_vals = []
            current_engaging = None
            engaging_col_positions = None

    return rows


def parse_code_to_systems_map(pdf_text):
    """Parse `COMPATIBLE WITH XXXX` blocks and extract listed implant systems."""
    code_to_systems = {}
    current_code = ""
    collect_systems = False
    expect_code_next_line = False

    for raw_line in pdf_text.splitlines():
        line = raw_line.strip()
        if not line:
            if collect_systems:
                collect_systems = False
            continue

        compatible = re.search(
            r"COMPATIBLE\s+WITH\s*([0-9]{4}[A-Z]?)", line, flags=re.IGNORECASE
        )
        if compatible:
            current_code = compatible.group(1).upper()
            code_to_systems.setdefault(current_code, [])
            collect_systems = False
            continue

        if re.fullmatch(r"COMPATIBLE\s+WITH", line, flags=re.IGNORECASE):
            expect_code_next_line = True
            collect_systems = False
            continue

        if expect_code_next_line and re.fullmatch(
            r"\d{4}[A-Z]?", line, flags=re.IGNORECASE
        ):
            current_code = line.upper()
            code_to_systems.setdefault(current_code, [])
            expect_code_next_line = False
            continue

        if expect_code_next_line:
            expect_code_next_line = False

        if "LIST OF COMPATIBILITIES AVAILABLE" in line.upper():
            collect_systems = True
            continue

        if not collect_systems or not current_code:
            continue

        if any(
            marker in line.upper()
            for marker in (
                "STANDARD DYNAMIC",
                "DYNAMIC 3TIBASE",
                "SCANBODY",
                "DYNAMIC MILLING TOOL",
                "SCREWS",
                "STRAIGHT",
                "MULTI-UNIT",
                "LIBRARY OPTIONS",
                "PRODUCT REFERENCES",
            )
        ):
            collect_systems = False
            continue

        candidates = [
            s.strip(" -") for s in re.split(r"\s+-\s+", line) if s.strip(" -")
        ]
        if len(candidates) == 1 and "-" in line and " - " not in line:
            candidates = [s.strip(" -") for s in line.split("-") if s.strip(" -")]

        for candidate in candidates:
            if len(candidate) < 2:
                continue
            if re.search(r"\d{2}\.\d{3}\.\d{3}\.\d{2}-\d", candidate):
                continue
            if candidate not in code_to_systems[current_code]:
                code_to_systems[current_code].append(candidate)

    return code_to_systems


def write_compatibility_options_csv(rows):
    """Write parsed PDF compatibility/option rows to CSV for auditing and import."""
    with open(COMPATIBILITY_OPTIONS_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "compatibility_code",
                "section",
                "reference",
                "reference_prefix",
                "reference_group",
                "reference_family",
                "reference_option",
                "reference_variant",
                "options",
                "source_line",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row["compatibility_code"],
                    row["section"],
                    row["reference"],
                    row["reference_prefix"],
                    row["reference_group"],
                    row["reference_family"],
                    row["reference_option"],
                    row["reference_variant"],
                    row["options"],
                    row["source_line"],
                ]
            )

    print(f"compatibility_options.csv: {len(rows)} rows → {COMPATIBILITY_OPTIONS_CSV}")


def build_option_map(parsed_rows):
    """Map exact reference SKU to merged option token strings from source rows."""
    by_ref: dict[str, list[str]] = {}
    for row in parsed_rows:
        ref = row.get("reference")
        if not ref:
            continue
        token = row.get("options", "")
        if token:
            by_ref.setdefault(ref, [])
            for part in token.split("|"):
                _merge_option_token(by_ref[ref], part)
    return {ref: "|".join(tokens) for ref, tokens in by_ref.items()}


def _merge_option_rows(option_map: dict[str, str], parsed_rows) -> dict[str, str]:
    """Merge additional parsed option rows into an existing reference option map."""
    for row in parsed_rows:
        ref = row.get("reference")
        options = row.get("options", "")
        if not ref or not options:
            continue
        tokens = [token for token in option_map.get(ref, "").split("|") if token]
        for part in options.split("|"):
            _merge_option_token(tokens, part)
        if ref.startswith("40.") and "TOTAL LENGTH(mm):" in options:
            tokens = [token for token in tokens if not token.startswith("LENGTH:")]
        option_map[ref] = "|".join(tokens)
    return option_map


def _merge_option_token(tokens: list[str], token: str) -> None:
    token = str(token or "").strip()
    if not token or ":" not in token:
        return

    key, value = [part.strip() for part in token.split(":", 1)]
    if not key or not value:
        return

    for index, existing in enumerate(tokens):
        if ":" not in existing:
            continue
        existing_key, existing_value = [part.strip() for part in existing.split(":", 1)]
        if existing_key != key:
            continue
        merged_values = []
        for item in re.split(r"/", existing_value):
            if item and item not in merged_values:
                merged_values.append(item)
        for item in re.split(r"/", value):
            if item and item not in merged_values:
                merged_values.append(item)
        if key in {"H(mm)", "GH(mm)", "LENGTH"}:
            merged_values = sorted(
                merged_values,
                key=lambda item: (
                    float(item.replace(",", "."))
                    if re.fullmatch(r"\d+(?:[\.,]\d+)?", item)
                    else 9999
                ),
            )
        tokens[index] = f"{key}:{'/'.join(merged_values)}"
        return

    tokens.append(f"{key}:{value}")


def _normalize_spec_number(value: str) -> str:
    return value.strip().replace(",", ".")


def _parse_dynamic_screw_spec_line(line: str) -> dict | None:
    m = re.match(
        r"^\s*(41\.\d{3}\.\d{3}\.\d{2}-\d)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\d+)\s*N[·.]?cm\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+"
        r"(\d+(?:[,.]\d+)?|-)\s+(\d+(?:[,.]\d+)?|-)\s+(\d+(?:[,.]\d+)?|-)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(straight|conical)\s+"
        r"(-|\d+\s*[º°])\s+(.+?)\s{2,}(Hexalobular.+?)\s*$",
        line,
        re.IGNORECASE,
    )
    if not m:
        return None

    (
        ref,
        metric,
        torque,
        total_length,
        thread_length,
        a_length,
        b_length,
        c_length,
        head_diameter,
        seat,
        angle,
        thread_entry,
        connection,
    ) = m.groups()

    specs = [
        ("METRIC", metric),
        ("TORQUE", f"{torque} N·cm"),
        ("TOTAL LENGTH(mm)", total_length),
        ("THREAD LENGTH(mm)", thread_length),
        ("A LENGTH(mm)", a_length),
        ("B LENGTH(mm)", b_length),
        ("C LENGTH(mm)", c_length),
        ("HEAD Ø(mm)", head_diameter),
        ("SEAT", seat.lower()),
        ("ANGLE", angle.replace(" ", "")),
        ("THREAD ENTRY", thread_entry.strip()),
        ("CONNECTION", connection.strip()),
    ]
    options = "|".join(
        f"{key}:{_normalize_spec_number(value) if key != 'TORQUE' else value}"
        for key, value in specs
        if value and value != "-"
    )
    return {"reference": ref, "options": options}


def _parse_straight_screw_spec_line(line: str) -> dict | None:
    m = re.match(
        r"^\s*(40\.\d{3}\.\d{3}\.\d{2}-\d)\s+"
        r"([A-Z]?\d+(?:-\d+|[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+(\d+\s*N[·.]?cm)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+"
        r"(\d+(?:[,.]\d+)?|-)\s+(\d+(?:[,.]\d+)?|-)\s+(\d+(?:[,.]\d+)?|-)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\S+)\s+(\S+)\s+(.+?)\s{2,}(.+?)\s*$",
        line,
        re.IGNORECASE,
    )
    if not m:
        return None

    (
        ref,
        metric,
        torque,
        total_length,
        thread_length,
        a_length,
        b_length,
        c_length,
        head_diameter,
        seat,
        angle,
        thread_entry_prefix,
        thread_entry,
        connection,
    ) = m.groups()
    thread_entry = " ".join(
        part.strip() for part in (thread_entry_prefix, thread_entry) if part.strip()
    )
    torque = _normalize_spec_number(torque)

    specs = [
        ("METRIC", metric),
        ("TORQUE", torque),
        ("TOTAL LENGTH(mm)", total_length),
        ("THREAD LENGTH(mm)", thread_length),
        ("A LENGTH(mm)", a_length),
        ("B LENGTH(mm)", b_length),
        ("C LENGTH(mm)", c_length),
        ("HEAD Ø(mm)", head_diameter),
        ("SEAT", seat),
        ("ANGLE", angle.strip()),
        ("THREAD ENTRY", thread_entry.strip()),
        ("CONNECTION", connection.strip()),
    ]
    options = "|".join(
        f"{key}:{_normalize_spec_number(value) if key != 'TORQUE' else value}"
        for key, value in specs
        if value and value != "-"
    )
    return {"reference": ref, "options": options}


def _parse_dynamic_milling_tool_spec_line(line: str) -> dict | None:
    m = re.match(
        r"^\s*.*?(33\.\d{3}\.\d{3}\.\d{2}-\d)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+"
        r"(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+(\d+(?:[,.]\d+)?)\s+"
        r"(\d+(?:[,.]\d+)?)\s*$",
        line,
    )
    if not m:
        return None

    (
        ref,
        cutting_diameter,
        seat,
        cutting_length,
        useful_length,
        stem_cutting_diameter,
        shank,
        total_length,
    ) = m.groups()
    specs = [
        ("CUTTING Ø(mm)", cutting_diameter),
        ("SEAT", seat),
        ("CUTTING LENGTH(mm)", cutting_length),
        ("USEFUL LENGTH(mm)", useful_length),
        ("STEM CUTTING Ø(mm)", stem_cutting_diameter),
        ("SHANK", shank),
        ("TOTAL LENGTH(mm)", total_length),
    ]
    options = "|".join(
        f"{key}:{_normalize_spec_number(value)}" for key, value in specs if value
    )
    return {"reference": ref, "options": options}


def parse_pdf_technical_spec_options(pdf_text: str) -> list[dict]:
    """Parse product technical-spec tables that are separate from compatibility rows."""
    rows: list[dict] = []
    section = None
    for raw_line in pdf_text.splitlines():
        line = raw_line.strip()
        if "DYNAMIC SCREWS TECHNICAL SPECIFICATIONS" in line:
            section = "dynamic_screws"
            continue
        if "SCREWS TECHNICAL SPECIFICATIONS" in line and "DYNAMIC" not in line:
            section = "straight_screws"
            continue
        if "DYNAMIC MILLING TOOL SPECIFICATIONS" in line:
            section = "dynamic_milling_tools"
            continue

        parsed = None
        if section == "dynamic_screws":
            parsed = _parse_dynamic_screw_spec_line(raw_line)
        elif section == "straight_screws":
            parsed = _parse_straight_screw_spec_line(raw_line)
        elif section == "dynamic_milling_tools":
            parsed = _parse_dynamic_milling_tool_spec_line(raw_line)
        if parsed:
            rows.append(parsed)

    return rows


def build_engaging_map(parsed_rows):
    """Map reference SKU to engaging value (0=non-engaging, 1=engaging, absent=unknown)."""
    by_ref = {}
    for row in parsed_rows:
        ref = row.get("reference")
        engaging = row.get("engaging")
        if ref and ref not in by_ref and engaging is not None:
            by_ref[ref] = engaging
    return by_ref


def build_price_lookup():
    """Build exact + wildcard pricing lookups from retail prices CSV."""
    exact = {}
    patterns = []

    with open(RETAIL_PRICES_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ref = row["reference"].strip()
            price = row["price_eur"].strip()
            if not ref:
                continue

            payload = {
                "name": row["name"].strip(),
                "price": price,
                "section": row["section"].strip(),
                "detail": row["detail"].strip(),
                "reference": ref,
            }
            norm = normalize_ref(ref)
            if re.search(r"[xXyY]", ref):
                patterns.append((ref_to_regex(ref), payload))
            else:
                exact[norm] = payload

    return exact, patterns


def match_price(reference_num, exact, patterns):
    """Match variant reference number to exact or wildcard retail pricing."""
    norm = normalize_ref(reference_num)
    if norm in exact:
        return exact[norm], "exact"

    for regex, payload in patterns:
        if regex.match(norm):
            return payload, "wildcard"

    return {"price": "", "section": "", "detail": "", "reference": ""}, "none"


_PRIMARY_CATALOG_SECTIONS = (
    "STANDARD DYNAMIC TIBASE",
    "DYNAMIC 3TIBASE",
    "DYNAMIC SCANBODY (LAB/CLIN)",
    "DYNAMIC MILLING TOOL",
    "DYNAMIC SCREWDRIVERS",
    "ANALOG",
    "STRAIGHT MULTI-UNIT",
    "STRAIGHT",
    "INTERNAL MULTI-UNIT",
    "MULTI-UNIT",
    "SCANBODY OP",
    "SCANBODY",
    "DAS MU SYSTEM COMPONENTS",
    "COMPLEMENTS",
    "DIRECTO IMPLANTE",
)


def _build_ref_lookups(code_to_systems):
    """Build {ref: [systems]} and {ref: section} from compatibility_options.csv.

    Used to assign categories to products whose segment_3 doesn't match any
    'COMPATIBLE WITH XXXX' block in the PDF (e.g. universal accessories).
    """
    ref_to_systems: dict[str, list[str]] = {}
    ref_to_section: dict[str, str] = {}
    section_priority = {s: i for i, s in enumerate(_PRIMARY_CATALOG_SECTIONS)}

    if not os.path.exists(COMPATIBILITY_OPTIONS_CSV):
        return ref_to_systems, ref_to_section

    with open(COMPATIBILITY_OPTIONS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ref = row.get("reference", "").strip()
            code = row.get("compatibility_code", "").strip()
            section = row.get("section", "").strip()
            if not ref or not code:
                continue
            for system in code_to_systems.get(code, []):
                if system not in ref_to_systems.get(ref, []):
                    ref_to_systems.setdefault(ref, []).append(system)
            # Keep the highest-priority (lowest index) section seen for this ref
            current = ref_to_section.get(ref, "")
            current_pri = section_priority.get(current, len(_PRIMARY_CATALOG_SECTIONS))
            new_pri = section_priority.get(section, len(_PRIMARY_CATALOG_SECTIONS))
            if not current or new_pri < current_pri:
                ref_to_section[ref] = section

    return ref_to_systems, ref_to_section


def build_merged_import_csv(
    option_map,
    code_to_systems,
    active_categories,
    engaging_map=None,
    product_type_map=None,
):
    """Merge products, prices and parsed PDF option families into one import-ready CSV."""
    exact, patterns = build_price_lookup()
    ref_to_systems, ref_to_section = _build_ref_lookups(code_to_systems)

    # Collect all product rows: main catalog first, then stubs (if file exists).
    # Yields (row, is_stub) so stubs can be forced hidden.
    def _iter_product_rows():
        with open(PRODUCTS_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                yield row, False
        if os.path.exists(PRODUCT_STUBS_CSV):
            with open(PRODUCT_STUBS_CSV, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    yield row, True

    count = 0
    with open(MERGED_IMPORT_CSV, "w", newline="", encoding="utf-8") as dst:
        writer = csv.writer(dst)
        writer.writerow(
            [
                "name",
                "description",
                "category",
                "price",
                "stock_quantity",
                "reference",
                "reference_num",
                "ean13",
                "price_match_type",
                "price_reference",
                "retail_name",
                "compatibility_code",
                "ref_segment_1",
                "ref_segment_2",
                "ref_segment_3",
                "ref_segment_4",
                "ref_check_digit",
                "options",
                "generated_description",
                "system_categories",
                "active_system_categories",
                "primary_system_category",
                "is_active_from_categories",
                "engaging",
                "catalog_section",
            ]
        )

        for row, is_stub in _iter_product_rows():
            ref = row["reference"].strip()
            ref_num = row["reference_num"].strip()
            name = row["name"].strip()
            parts = parse_reference_parts(ref)
            price_payload, match_type = match_price(ref_num, exact, patterns)
            family_code = parts["segment_3"]
            padded_family = family_code.zfill(4) if family_code else ""
            # Use only explicit PDF listing (compatibility_options.csv).
            # segment_3-based heuristic was removed: a product must appear in a
            # PDF "COMPATIBLE WITH XXXX" section to receive compatibility systems.
            systems = ref_to_systems.get(ref, [])
            active_systems = [
                active_categories[normalize_system_name(system)]
                for system in systems
                if normalize_system_name(system) in active_categories
            ]
            primary_system = (
                active_systems[0] if active_systems else (systems[0] if systems else "")
            )
            is_active = 1 if active_systems else 0
            catalog_section = (product_type_map or {}).get(ref) or ref_to_section.get(
                ref, ""
            )
            if name.upper().startswith("DMT "):
                catalog_section = "DYNAMIC MILLING TOOL"
            retail_name = price_payload.get("name", "")
            row_tokens = []
            for part in option_map.get(ref, "").split("|"):
                _merge_option_token(row_tokens, part)
            _merge_product_name_option_tokens(
                row_tokens,
                _product_name_option_tokens(name, reference=ref),
            )
            row_tokens, catalog_section = _apply_product_manual_override(
                ref, row_tokens, catalog_section
            )
            option_tokens = "|".join(row_tokens)

            description_parts = [
                price_payload["detail"],
                f"Parametre: {option_tokens}" if option_tokens else "",
            ]
            generated_description = " | ".join([p for p in description_parts if p])
            engaging_val = (engaging_map or {}).get(ref, "")
            engaging_str = "" if engaging_val == "" else str(engaging_val)

            writer.writerow(
                [
                    name,
                    price_payload["detail"],
                    price_payload["section"],
                    price_payload["price"],
                    0,
                    ref,
                    ref_num,
                    row.get("ean13", "").strip(),
                    match_type,
                    price_payload["reference"],
                    retail_name,
                    padded_family,
                    parts["segment_1"],
                    parts["segment_2"],
                    parts["segment_3"],
                    parts["segment_4"],
                    parts["check_digit"],
                    option_tokens,
                    generated_description,
                    ";".join(systems),
                    ";".join(active_systems),
                    primary_system,
                    is_active,
                    engaging_str,
                    catalog_section,
                ]
            )
            count += 1

    print(f"import_all_merged.csv: {count} rows → {MERGED_IMPORT_CSV}")


# Known OCR/parsing errors in PDF system names → correct canonical name
_SYSTEM_NAME_CORRECTIONS = {
    "ANKLYOS": "ANKYLOS",
    "BIOMET 3L": "BIOMET 3i",
    "MEDINTAKA": "MEDENTIKA",
}


def _apply_system_name_corrections(code_to_systems):
    """Fix known misspellings produced by PDF text extraction."""
    return {
        code: [_SYSTEM_NAME_CORRECTIONS.get(s, s) for s in systems]
        for code, systems in code_to_systems.items()
    }


def _row_to_option_tokens(row: dict) -> str:
    """Convert a parse_catalog.py structured product row to pipe-separated key:value tokens."""
    parts = []
    product_type = row.get("product_type")
    if product_type in {"ADAPTOR", "SCREWDRIVER ADAPTOR", "DYNAMIC MILLING TOOL"}:
        parts.append(f"TYPE:{product_type}")
    if row.get("GH_mm"):
        parts.append(f"GH(mm):{row['GH_mm']}")
    alpha_s = row.get("alpha_S")
    if alpha_s:
        if "CH=" in str(alpha_s):
            for segment in str(alpha_s).split(" / "):
                m = re.match(r"CH=(\d+)mm:(.+)", segment.strip())
                if m:
                    parts.append(f"αS(CH={m.group(1)}mm):{m.group(2)}")
        else:
            parts.append(f"αS:{alpha_s}")
    if row.get("alpha_C"):
        parts.append(f"αC:{row['alpha_C']}")
    if row.get("alpha_di"):
        parts.append(f"αdi:{row['alpha_di']}")
    if row.get("H_mm"):
        h_values = str(row["H_mm"]).split("/")
        if row.get("product_type") in {"ADAPTOR", "SCREWDRIVER", "SCREWDRIVER ADAPTOR"}:
            h_values = [value for value in h_values if value in {"8", "10", "12"}]
        if h_values:
            parts.append(f"H(mm):{'/'.join(h_values)}")
    if row.get("length_mm"):
        parts.append(f"LENGTH:{row['length_mm']}")
        if row.get("product_type") == "DYNAMIC SCREW":
            parts.append(f"DYNAMIC SCREW:{row.get('sku') or row['length_mm']}")
    if row.get("height_mm"):
        parts.append(f"H(mm):{row['height_mm']}")
    if row.get("type_label"):
        if row.get("product_type") in {
            "DYNAMIC MILLING TOOL",
            "ANALOG",
            "DIGITAL ANALOG",
        }:
            parts.append(f"SHANK:{row['type_label']}")
        elif "SCREW" in str(row.get("product_type", "")):
            parts.append(f"TYPE:{row['type_label']}")
        else:
            parts.append(f"Typ:{row['type_label']}")
    if row.get("screwdriver_refs"):
        parts.append(f"SCREWDRIVER:{row['screwdriver_refs']}")
    if row.get("straight_screw_refs"):
        parts.append(f"STRAIGHT SCREW:{row['straight_screw_refs']}")
    if row.get("dynamic_screw_refs"):
        parts.append(f"DYNAMIC SCREW:{row['dynamic_screw_refs']}")
    if row.get("adaptor_refs"):
        parts.append(f"ADAPTOR:{row['adaptor_refs']}")
    if row.get("screwdriver_adaptor_refs"):
        parts.append(f"SCREWDRIVER ADAPTOR:{row['screwdriver_adaptor_refs']}")
    return "|".join(parts[:8])


_PRODUCT_ALPHA_DI_CORRECTIONS = {
    "33.320.704.01-2": "25",
    "33.420.704.01-2": "25",
    "33.620.704.01-2": "25",
    "33.435.758.01-2": "30",
}


_CAPS_REFERENCE_OPTIONS = {
    "49.418.000.01-2": ("3.8", "Regular"),
    "49.418.000.02-2": ("3.8", "Wide"),
    "49.419.000.01-2": ("6", "Regular"),
    "49.419.000.02-2": ("6", "Wide"),
    "49.420.000.01-2": ("8", "Regular"),
    "49.420.000.02-2": ("8", "Wide"),
}


def _product_name_option_tokens(name: str, reference: str = "") -> str:
    """Extract missing technical dimensions from product names when the PDF table omits them."""
    parts = []
    caps_options = _CAPS_REFERENCE_OPTIONS.get(reference)
    if caps_options:
        height, type_label = caps_options
        parts.append(f"H(mm):{height}")
        parts.append(f"Typ:{type_label}")

    das_multi_unit_m = re.search(
        r"\bDAS\s+Multi-Unit\s+(\d+(?:[\.,]\d+)?)\s*[º°].*?\bG(\d+(?:[\.,]\d+)?(?:/\d+(?:[\.,]\d+)?)?)",
        name,
        re.IGNORECASE,
    )
    if das_multi_unit_m:
        angle = das_multi_unit_m.group(1).replace(",", ".")
        gh = das_multi_unit_m.group(2).replace(",", ".")
        parts.append(f"GH(mm):{gh}")
        parts.append(f"TYPE:ANGULATED MULTI-UNIT {angle}º")

    seat_m = re.search(r"\bSeat\s+(\d+(?:[\.,]\d+)?)\s*[º°]", name, re.IGNORECASE)
    alpha_di = _PRODUCT_ALPHA_DI_CORRECTIONS.get(reference)
    if not alpha_di and seat_m:
        alpha_di = seat_m.group(1).replace(",", ".")
    if alpha_di:
        parts.append(f"αdi:{alpha_di}°")
    shank_m = re.search(r"\bShank\s+(\d+(?:[\.,]\d+)?)\s*mm\b", name, re.IGNORECASE)
    if shank_m:
        parts.append(f"SHANK:{shank_m.group(1).replace(',', '.')}")
    return "|".join(parts)


def _merge_product_name_option_tokens(tokens: list[str], token_source: str) -> None:
    """Merge exact product-name tokens; row-specific GH replaces shared PDF GH."""
    for part in token_source.split("|"):
        if part.startswith("GH(mm):"):
            tokens[:] = [
                token
                for token in tokens
                if ":" not in token or token.split(":", 1)[0].strip() != "GH(mm)"
            ]
        _merge_option_token(tokens, part)


_PRODUCT_MANUAL_OVERRIDES = {
    # PDF/source-table overrides that cannot be inferred reliably from a single row.
    "41.320.117.01-2": {
        "remove_keys": {"DYNAMIC SCREW", "LENGTH", "SCREWDRIVER"},
        "set_tokens": [
            "DYNAMIC SCREW:41.320.075.01-2",
            "LENGTH:18/24/32",
            "SCREWDRIVER:43.618.201.01-2/43.624.201.01-2/43.632.201.01-2",
        ],
    },
    "43.618.201.01-2": {
        "catalog_section": "SCREWDRIVER",
        "remove_keys": {"H(mm)", "DYNAMIC SCREW"},
        "set_tokens": ["LENGTH:18", "DYNAMIC SCREW:41.316.080.01-2"],
    },
    "43.624.201.01-2": {
        "catalog_section": "SCREWDRIVER",
        "remove_keys": {"H(mm)"},
        "set_tokens": ["LENGTH:24"],
    },
    "43.632.201.01-2": {
        "catalog_section": "SCREWDRIVER",
        "remove_keys": {"H(mm)", "DYNAMIC SCREW"},
        "set_tokens": ["LENGTH:32", "DYNAMIC SCREW:41.316.094.01-2"],
    },
    "43.625.108.01-2": {
        "remove_keys": {"H(mm)"},
    },
    "23.313.025.01-2": {
        "remove_keys": {"H(mm)", "GH(mm)"},
    },
    "23.413.025.01-2": {
        "remove_keys": {"H(mm)", "GH(mm)"},
    },
    "53.413.025.01-2": {
        "remove_keys": {"H(mm)"},
    },
}


def _apply_product_manual_override(
    reference: str, row_tokens: list[str], catalog_section: str
) -> tuple[list[str], str]:
    override = _PRODUCT_MANUAL_OVERRIDES.get(reference)
    if not override:
        return row_tokens, catalog_section

    remove_keys = set(override.get("remove_keys", set()))
    tokens = [
        token
        for token in row_tokens
        if ":" in token and token.split(":", 1)[0].strip() not in remove_keys
    ]
    for token in override.get("set_tokens", []):
        _merge_option_token(tokens, token)

    return tokens, override.get("catalog_section", catalog_section)


def _build_product_type_map(catalog_rows: list[dict]) -> dict[str, str]:
    """Choose the storefront product type from structured PDF rows."""
    priority = {
        "DYNAMIC MILLING TOOL": 1,
        "ADAPTOR": 2,
        "SCREWDRIVER ADAPTOR": 3,
        "DYNAMIC SCANBODY": 4,
        "DYNAMIC SCREW": 5,
        "SCREWDRIVER": 6,
    }
    by_ref: dict[str, str] = {}
    best_priority: dict[str, int] = {}
    for row in catalog_rows:
        ref = row.get("sku")
        product_type = row.get("product_type")
        if not ref or not product_type:
            continue
        rank = priority.get(product_type, 100)
        if ref not in by_ref or rank < best_priority[ref]:
            by_ref[ref] = product_type
            best_priority[ref] = rank
    return by_ref


def convert_catalog_pdf_options():
    """Parse PDF catalog and generate compatibility options CSV + merged import CSV."""
    import sys

    pdf_path = resolve_source_file("PRODUCT-REFERENCE-0326_01.pdf")
    active_categories = load_active_categories(VISIBLE_CATEGORIES_TXT)

    # Use the layout-aware parse_catalog.py parser for structured, accurate data
    try:
        if BASE_DIR not in sys.path:
            sys.path.insert(0, BASE_DIR)
        from parse_catalog import (
            extract_text as _extract_text,
            parse_products as _parse_products,
        )

        product_text = _extract_text(pdf_path, first_page=43, last_page=329)
        catalog_rows = _parse_products(product_text, pdf_first_page=43)
        product_type_map = _build_product_type_map(catalog_rows)
        option_map = build_option_map(
            [
                {"reference": row["sku"], "options": _row_to_option_tokens(row)}
                for row in catalog_rows
                if row.get("sku")
            ]
        )
        engaging_map = {
            row["sku"]: row["engaging"]
            for row in catalog_rows
            if row.get("sku") and row.get("engaging") is not None
        }
        print(
            f"parse_catalog: {len(catalog_rows)} rows, {len(option_map)} option tokens, "
            f"{len(engaging_map)} engaging values"
        )
    except Exception as exc:
        print(f"Warning: parse_catalog failed ({exc}), falling back to legacy parser")
        pdf_text = load_pdf_text(pdf_path)
        if not pdf_text:
            print("compatibility_options.csv: skipped (could not parse PDF text)")
            build_merged_import_csv({}, {}, active_categories)
            return
        rows = parse_pdf_compatibility_rows(pdf_text)
        write_compatibility_options_csv(rows)
        option_map = build_option_map(rows)
        engaging_map = build_engaging_map(rows)
        pdf_text_for_systems = pdf_text
    else:
        pdf_text_for_systems = load_pdf_text(pdf_path)
        if pdf_text_for_systems:
            rows = parse_pdf_compatibility_rows(pdf_text_for_systems)
            write_compatibility_options_csv(rows)
        else:
            print("compatibility_options.csv: skipped (could not parse PDF text)")

    code_to_systems = (
        parse_code_to_systems_map(pdf_text_for_systems) if pdf_text_for_systems else {}
    )
    code_to_systems = _apply_system_name_corrections(code_to_systems)
    if pdf_text_for_systems:
        technical_spec_rows = parse_pdf_technical_spec_options(pdf_text_for_systems)
        _merge_option_rows(option_map, technical_spec_rows)
        print(f"technical specs: {len(technical_spec_rows)} option rows")
    build_merged_import_csv(
        option_map,
        code_to_systems,
        active_categories,
        engaging_map,
        product_type_map if "product_type_map" in locals() else None,
    )


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "dealer":
        print("Converting dealer prices...\n")
        convert_dealer_prices()
    else:
        print("Converting Excel files to CSV...\n")
        convert_products()
        convert_retail_prices()
        convert_dealer_prices()
        convert_catalog_pdf_options()
    print("\nDone. Files are in data/csv/ (gitignored).")
